import magic
import fitz
import zipfile
import tempfile
from pathlib import Path
from docx import Document
from openpyxl import load_workbook
PARSERS_MAPPING = {
    "application/pdf": "pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
    "application/msword": "docx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
    "application/vnd.ms-excel": "xlsx",
    "application/vnd.ms-excel.sheet.macroEnabled.12": "xlsx",
    "application/zip": "zip",
    "application/x-zip-compressed": "zip",
    "text/csv": "csv",
    "text/plain": "txt"
}
file_path = "beta_doc/C.4-Cahier_PT_SC.pdf"
file_docx = "beta_doc/4 MA Rénov plateau sportif.docx"
file_xlsx = "beta_doc/MAPA_2026_08 _ DPGF.xlsx"
file_zip = "beta_doc/DCE_test.zip"
def detecter_parser(chemin_fichier: str) -> str:
    detector = magic.Magic(mime=True)
    type_mime = detector.from_file(chemin_fichier)

    return PARSERS_MAPPING.get(type_mime, "PARSER_INCONNU")

def detecter_type_page(page):
    """
    Retourne 'native', 'scanned', ou 'hybrid'
    """
    texte = page.get_text().strip()
    nb_chars = len(texte)
    
    if nb_chars > 0:
        nb_garbled = texte.count("\ufffd") + texte.count("?")
        ratio_garbled = nb_garbled / nb_chars
    else:
        ratio_garbled = 0
    
    page_area = abs(page.rect)
    image_area = 0
    for img in page.get_images(full=True):
        for bbox in page.get_image_rects(img[0]):
            image_area += abs(bbox & page.rect)
    ratio_image = image_area / page_area if page_area > 0 else 0
    
    fonts = page.get_fonts()
    has_ocr_font = any("Glyphless" in str(f) for f in fonts)
    
    if nb_chars < 50 and ratio_image > 0.8:
        return 'scanned'           
    elif ratio_garbled > 0.3:
        return 'scanned'           
    elif has_ocr_font and nb_chars > 100:
        return 'native_ocr'       
    elif nb_chars > 100:
        return 'native'           
    else:
        return 'scanned'           
    

def type_page(chemin_fichier: str):
    with fitz.open(chemin_fichier) as doc:
        for page in doc:
            num_page = page.number + 1
            print(f"Page {num_page}, {detecter_type_page(page)}")
    
def extraire_texte_page(chemin_fichier: str):
    with fitz.open(chemin_fichier) as doc:
        text = []
        for page in doc:
            for bloc in page.get_text("blocks"):
                text.append({
                    "texte": bloc[4],
                    "page": page.number + 1,
                    "bbox": (bloc[0], bloc[1], bloc[2], bloc[3])
                })
    return text
def chunker(chemin_fichier: str) -> str:
    chunks = []
    blocs =  extraire_texte_page(chemin_fichier)
    for chunk in range(0, len(blocs), 5):
        groupe = blocs[chunk:chunk+5]
        chunks.append({
            "texte": " ".join([b["texte"] for b in groupe]),
            "page": groupe[0]["page"],
            "bbox": groupe[0]["bbox"]
        })
    return chunks
    
def extraire_docx(chemin_fichier: str) -> str:
    paragraphes = []
    doc = Document(chemin_fichier)
    for p in doc.paragraphs:
        paragraphes.append({"texte": p.text, "style": p.style.name})
    return paragraphes


def extraire_xlsx(chemin_fichier: str) -> str:
    wb = load_workbook(chemin_fichier, data_only=True)
    sheet = wb.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        if any(row):
            rows.append(row)
    wb.close()
    return rows

def extraire_zip(chemin_fichier: str) -> str:
    dossier_temp = tempfile.mkdtemp()
    with zipfile.ZipFile(chemin_fichier, "r") as archive:
        archive.extractall(path=dossier_temp)
        dossier = Path(dossier_temp)
    fichiers = [f for f in dossier.rglob("*") if f.is_file()]
    return fichiers

print(extraire_zip(file_zip))